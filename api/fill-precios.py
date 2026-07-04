# api/fill-precios.py
# Vercel Serverless Function (Python runtime)
# Lee el archivo de cambios de precios, rellena la plantilla elegida y la devuelve

from http.server import BaseHTTPRequestHandler
import json, base64, io, cgi, os

def fill_template(template_bytes, template_name, products):
    """
    Rellena la plantilla reemplazando DESCRIPCION y PRECIO con los datos reales.
    Conserva logos, imágenes, formato y estructura intactos.
    """
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet

    wb = load_workbook(
        io.BytesIO(template_bytes),
        read_only=False,
        data_only=False,
        keep_vba='xlsm' in template_name.lower()
    )

    prod_idx = 0  # índice del siguiente producto a colocar

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not isinstance(ws, Worksheet):
            continue

        # Encontrar todas las celdas DESCRIPCION y PRECIO en orden
        desc_cells  = []
        price_cells = []

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    continue
                vs = str(v).strip().upper()
                if vs == 'DESCRIPCION':
                    desc_cells.append((r, c))
                elif vs == 'PRECIO':
                    price_cells.append((r, c))

        # Cada par (desc_cell, price_cell) es un slot
        slots = list(zip(desc_cells, price_cells))

        for (dr, dc), (pr, pc) in slots:
            if prod_idx >= len(products):
                # Sin más productos — limpiar la celda
                ws.cell(row=dr, column=dc).value = ''
                ws.cell(row=pr, column=pc).value = ''
            else:
                p = products[prod_idx]
                # Descripción
                ws.cell(row=dr, column=dc).value = p['name']
                # Precio en formato moneda: $489.00
                price_val = p['price']
                ws.cell(row=pr, column=pc).value = price_val
                # Formato de número moneda con 2 decimales
                ws.cell(row=pr, column=pc).number_format = '"$"#,##0.00'
                prod_idx += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def parse_cambios(file_bytes):
    """
    Lee el archivo de cambios de precios.
    Última hoja, col B = descripción, col D = precio, desde fila 3.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = wb.sheetnames[-1]  # última hoja
    ws = wb[sheet_name]

    products = []
    for r in range(3, (ws.max_row or 200) + 1):
        name  = ws.cell(row=r, column=2).value  # col B
        price = ws.cell(row=r, column=4).value  # col D
        if name is None or str(name).strip() == '':
            continue
        name  = str(name).strip().upper()
        price = float(str(price).replace(',', '.').replace('$', '').strip()) if price is not None else None
        if price is None:
            continue
        products.append({'name': name, 'price': price})

    return products, sheet_name


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        self._cors()
        content_type = self.headers.get('Content-Type', '')

        try:
            # Leer body JSON con los archivos en base64
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))

            cambios_b64  = body['cambios']     # base64 del archivo de cambios
            template_key = body['template']    # 'iman' o 'sin_iman'

            # Decodificar archivo de cambios
            cambios_bytes = base64.b64decode(cambios_b64)

            # Leer productos del archivo de cambios
            products, sheet_name = parse_cambios(cambios_bytes)

            if not products:
                self._json(400, {'error': f'No se encontraron productos en la hoja "{sheet_name}"'})
                return

            # Cargar plantilla desde carpeta /plantillas
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            templates = {
                'iman':     os.path.join(base_dir, 'plantillas', 'PRECIOS_IMAN.xlsm'),
                'sin_iman': os.path.join(base_dir, 'plantillas', 'PRECIOS_SIN_IMAN.xlsx'),
            }

            if template_key not in templates:
                self._json(400, {'error': f'Plantilla desconocida: {template_key}'})
                return

            template_path = templates[template_key]
            with open(template_path, 'rb') as f:
                template_bytes = f.read()

            template_name = os.path.basename(template_path)

            # Rellenar plantilla
            result_bytes = fill_template(template_bytes, template_name, products)

            # Responder con el archivo rellenado en base64
            result_b64 = base64.b64encode(result_bytes).decode()
            ext = 'xlsm' if template_key == 'iman' else 'xlsx'
            filename = f'PRECIOS_{template_key.upper()}_{sheet_name.replace(" ","-")}.{ext}'

            self._json(200, {
                'ok': True,
                'file': result_b64,
                'filename': filename,
                'count': len(products),
                'sheet': sheet_name,
                'ext': ext,
            })

        except Exception as e:
            import traceback
            self._json(500, {'error': str(e), 'trace': traceback.format_exc()})

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _json(self, status, data):
        body = json.dumps(data).encode()
        self.send_response(status)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)
